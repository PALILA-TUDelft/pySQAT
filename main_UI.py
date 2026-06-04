# main_UI.py
# SQAT – Primary PySide6 GUI for SQAT4PY

import sys
import os
import time
import threading
import queue
from pathlib import Path
from typing import Any, Dict, Optional, List

import numpy as np

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QListWidget, QListWidgetItem, QComboBox,
    QCheckBox, QLineEdit, QTextEdit, QProgressBar, QFrame,
    QSplitter, QFileDialog, QMessageBox, QAbstractItemView,
    QSizePolicy, QGroupBox, QGridLayout, QDialog, QStatusBar,
)
from PySide6.QtCore import Qt, QTimer, QSize, Signal, QObject
from PySide6.QtGui import QFont, QColor, QTextCursor, QPalette, QPixmap

import matplotlib
matplotlib.use("QtAgg")   # must be called before pyplot/backend imports
FigureCanvasQTAgg = None
NavigationToolbar2QT = None


# ── heavy imports are deferred to first use so the UI opens instantly ────────
# openpyxl, metrics_*, utilities, and plotting_metrics are loaded lazily.
# Only the bare minimum is imported at module level.

def _lazy_imports():
    """Load all heavy modules; called once before first analysis or graph window."""
    global FigureCanvasQTAgg, NavigationToolbar2QT
    global wav2sig, see
    global Loudness_ISO532_1, EPNL_FAR_Part36
    global Sharpness_DIN45692
    global Roughness_Daniel1997
    global FluctuationStrength_Osses2016
    global Tonality_Aures1985
    global PsychoacousticAnnoyance_Di2016, PsychoacousticAnnoyance_Zwicker1999, PsychoacousticAnnoyance_More2010
    global create_metric_plot, create_metric_plot_split
    global Workbook, get_column_letter, XFont, PatternFill, XAlignment, Border, Side

    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as _fca, NavigationToolbar2QT as _ntb
    FigureCanvasQTAgg, NavigationToolbar2QT = _fca, _ntb

    from utilities import wav2sig as _w, see as _s
    wav2sig, see = _w, _s

    from metrics_loudness import Loudness_ISO532_1 as _a
    from metrics_epnl import EPNL_FAR_Part36 as _b
    Loudness_ISO532_1, EPNL_FAR_Part36 = _a, _b

    from metrics_sharpness import Sharpness_DIN45692 as _c
    Sharpness_DIN45692 = _c

    from metrics_roughness import Roughness_Daniel1997 as _d
    Roughness_Daniel1997 = _d

    from metrics_fluctuation import FluctuationStrength_Osses2016 as _e
    FluctuationStrength_Osses2016 = _e

    from metrics_tonality import Tonality_Aures1985 as _f
    Tonality_Aures1985 = _f

    from metrics_annoyance import (
        PsychoacousticAnnoyance_Di2016 as _g,
        PsychoacousticAnnoyance_Zwicker1999 as _h,
        PsychoacousticAnnoyance_More2010 as _i,
    )
    PsychoacousticAnnoyance_Di2016 = _g
    PsychoacousticAnnoyance_Zwicker1999 = _h
    PsychoacousticAnnoyance_More2010 = _i

    from plotting_metrics import create_metric_plot as _j, create_metric_plot_split as _k
    create_metric_plot, create_metric_plot_split = _j, _k

    from openpyxl import Workbook as _W
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.styles import Font as _XF, PatternFill as _PF, Alignment as _XA, Border as _Bd, Side as _Sd
    Workbook, get_column_letter = _W, _gcl
    XFont, PatternFill, XAlignment, Border, Side = _XF, _PF, _XA, _Bd, _Sd


# Placeholders so names exist at module level before lazy load
wav2sig = see = None
Loudness_ISO532_1 = EPNL_FAR_Part36 = None
Sharpness_DIN45692 = Roughness_Daniel1997 = FluctuationStrength_Osses2016 = None
Tonality_Aures1985 = None
PsychoacousticAnnoyance_Di2016 = PsychoacousticAnnoyance_Zwicker1999 = PsychoacousticAnnoyance_More2010 = None
create_metric_plot = create_metric_plot_split = None
Workbook = get_column_letter = None
XFont = PatternFill = XAlignment = Border = Side = None

_imports_done = False


# ── application constants ───────────────────────────────────────────────────

METRICS = [
    "Loudness (ISO 532-1)",
    "Loudness (ECMA 418-2)",
    "Sharpness (DIN 45692)",
    "Roughness (Daniel 1997)",
    "Roughness (ECMA 418-2)",
    "Fluctuation Strength (Osses 2016)",
    "Tonality (Aures 1985)",
    "Tonality (ECMA 418-2)",
    "Annoyance (Di 2016)",
    "Annoyance (Zwicker 1999)",
    "Annoyance (More 2010)",
    "EPNL (FAR Part 36)",
]

DEFAULTS = {
    "Loudness (ISO 532-1)":              dict(field=0, method=2, time_skip=0.5),
    "Loudness (ECMA 418-2)":             dict(field=0, method=2, time_skip=0.5),
    "Sharpness (DIN 45692)":             dict(field=0, method=2, weight_type="DIN45692", time_skip=0.5),
    "Roughness (Daniel 1997)":           dict(time_skip=0.0),
    "Roughness (ECMA 418-2)":            dict(time_skip=0.5),
    "Fluctuation Strength (Osses 2016)": dict(method_fs=1, time_skip=0.0),
    "Tonality (Aures 1985)":             dict(field=0, time_skip=0.0),
    "Tonality (ECMA 418-2)":             dict(field=0, time_skip=0.5),
    "Annoyance (Di 2016)":               dict(field=0, time_skip=0.2),
    "Annoyance (Zwicker 1999)":          dict(field=0, time_skip=0.2),
    "Annoyance (More 2010)":             dict(field=0, time_skip=0.2),
    "EPNL (FAR Part 36)":                dict(method_epnl=1, threshold_epnl=None),
}

PARAM_SPECS = {
    "time_skip":      {"label": "Time skip [s]",                       "kind": "entry",
                       "options": []},
    "field":          {"label": "Sound field",                         "kind": "combo",
                       "options": [("0 = free field", 0), ("1 = diffuse field", 1)]},
    "method":         {"label": "Method",                              "kind": "combo",
                       "options": [("1 = stationary", 1), ("2 = time-varying", 2)]},
    "weight_type":    {"label": "Sharpness weight",                    "kind": "combo",
                       "options": [("DIN45692","DIN45692"), ("aures","aures"), ("bismarck","bismarck")]},
    "method_fs":      {"label": "FS method",                           "kind": "combo",
                       "options": [("0 = stationary", 0), ("1 = time-varying", 1)]},
    "threshold_epnl": {"label": "EPNL tone-threshold [PNdB] (opt.)",  "kind": "entry",
                       "options": []},
}

METRIC_PARAM_LAYOUT = {
    "Loudness (ISO 532-1)":              ["field", "method", "time_skip"],
    "Loudness (ECMA 418-2)":             ["field", "method", "time_skip"],
    "Sharpness (DIN 45692)":             ["field", "method", "weight_type", "time_skip"],
    "Roughness (Daniel 1997)":           ["time_skip"],
    "Roughness (ECMA 418-2)":            ["field", "method", "time_skip"],
    "Fluctuation Strength (Osses 2016)": ["method_fs", "time_skip"],
    "Tonality (Aures 1985)":             ["field", "time_skip"],
    "Tonality (ECMA 418-2)":             ["field", "method", "time_skip"],
    "Annoyance (Di 2016)":               ["field", "time_skip"],
    "Annoyance (Zwicker 1999)":          ["field", "time_skip"],
    "Annoyance (More 2010)":             ["field", "time_skip"],
    "EPNL (FAR Part 36)":                ["threshold_epnl"],
}


# ── colour palettes ─────────────────────────────────────────────────────────

DARK_COLORS: Dict[str, str] = {
    "bg":       "#0f172a",
    "surface":  "#1e293b",
    "card":     "#263348",
    "border":   "#334155",
    "text":     "#f1f5f9",
    "text2":    "#94a3b8",
    "text3":    "#64748b",
    "accent":   "#6366f1",
    "accent2":  "#818cf8",
    "success":  "#10b981",
    "success2": "#34d399",
    "danger":   "#ef4444",
    "warning":  "#f59e0b",
    "header":   "#0a0f1e",
    "console_bg":   "#060d1a",
    "console_text": "#7dd3a8",
}

LIGHT_COLORS: Dict[str, str] = {
    "bg":       "#f1f5f9",
    "surface":  "#ffffff",
    "card":     "#e8eef5",
    "border":   "#cbd5e1",
    "text":     "#1e293b",
    "text2":    "#475569",
    "text3":    "#94a3b8",
    "accent":   "#6366f1",
    "accent2":  "#4f46e5",
    "success":  "#059669",
    "success2": "#10b981",
    "danger":   "#dc2626",
    "warning":  "#d97706",
    "header":   "#ffffff",
    "console_bg":   "#f8fafc",
    "console_text": "#065f46",
}

C: Dict[str, str] = dict(DARK_COLORS)


def make_style(C: Dict[str, str]) -> str:
    return f"""
QMainWindow, QDialog {{
    background: {C['bg']};
}}
QWidget {{
    background: {C['bg']};
    color: {C['text']};
    font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    font-size: 13px;
    selection-background-color: {C['accent']};
    selection-color: #ffffff;
}}

/* ── scrollbars ── */
QScrollBar:vertical {{
    background: {C['surface']};
    width: 8px;
    border-radius: 4px;
    margin: 0;
}}
QScrollBar::handle:vertical {{
    background: {C['border']};
    border-radius: 4px;
    min-height: 24px;
}}
QScrollBar::handle:vertical:hover {{ background: {C['text3']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: {C['surface']};
    height: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {C['border']};
    border-radius: 4px;
    min-width: 24px;
}}
QScrollBar::handle:horizontal:hover {{ background: {C['text3']}; }}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}

/* ── sidebar ── */
#sidebar {{
    background: {C['surface']};
    border-right: 1px solid {C['border']};
}}

/* ── metric list ── */
QListWidget {{
    background: {C['surface']};
    border: none;
    outline: none;
    padding: 4px;
}}
QListWidget::item {{
    padding: 10px 12px;
    border-radius: 6px;
    margin: 2px 4px;
    color: {C['text2']};
    border-left: 3px solid transparent;
}}
QListWidget::item:hover {{
    background: {C['card']};
    color: {C['text']};
}}
QListWidget::item:selected {{
    background: rgba(99,102,241,0.18);
    color: {C['accent2']};
    border-left: 3px solid {C['accent']};
}}

/* ── buttons ── */
QPushButton {{
    background: {C['card']};
    color: {C['text']};
    border: 1px solid {C['border']};
    border-radius: 8px;
    padding: 8px 18px;
    font-size: 13px;
    font-weight: 500;
}}
QPushButton:hover {{
    background: {C['border']};
    border-color: {C['text3']};
}}
QPushButton:pressed {{ background: {C['surface']}; }}
QPushButton:disabled {{
    color: {C['text3']};
    border-color: {C['surface']};
    background: {C['surface']};
}}

QPushButton#btn_run {{
    background: {C['success']};
    color: #ffffff;
    border: 2px solid {C['success2']};
    border-radius: 10px;
    padding: 12px 24px;
    font-size: 15px;
    font-weight: 700;
    letter-spacing: 0.5px;
}}
QPushButton#btn_run:hover  {{ background: {C['success2']}; border-color: {C['success']}; color: #ffffff; }}
QPushButton#btn_run:pressed {{ background: #059669; border-color: #047857; color: #ffffff; }}
QPushButton#btn_run:disabled {{ background: transparent; color: {C['text3']}; border: 2px solid {C['border']}; }}

#btn_graphs {{
    background: rgba(99,102,241,0.15);
    color: {C['accent2']};
    border: 1px solid rgba(99,102,241,0.5);
    border-radius: 8px;
    padding: 9px 18px;
    font-size: 13px;
    font-weight: 600;
}}
#btn_graphs:hover  {{ background: rgba(99,102,241,0.28); border-color: {C['accent']}; }}
#btn_graphs:disabled {{ background: transparent; color: {C['text3']}; border: 1px solid {C['border']}; }}

#btn_wave {{
    background: rgba(245,158,11,0.12);
    color: {C['warning']};
    border: 1px solid rgba(245,158,11,0.5);
    border-radius: 8px;
    padding: 9px 18px;
    font-size: 13px;
    font-weight: 600;
}}
#btn_wave:hover {{ background: rgba(245,158,11,0.22); border-color: {C['warning']}; }}
#btn_wave:disabled {{ background: transparent; color: {C['text3']}; border: 1px solid {C['border']}; }}

/* ── checkboxes ── */
QCheckBox {{
    spacing: 8px;
    color: {C['text2']};
    font-size: 13px;
}}
QCheckBox::indicator {{
    width: 18px;
    height: 18px;
    border-radius: 5px;
    border: 2px solid {C['border']};
    background: {C['surface']};
}}
QCheckBox::indicator:hover {{ border-color: {C['accent']}; }}
QCheckBox::indicator:checked {{
    background: {C['accent']};
    border: 2px solid {C['accent']};
}}

/* ── combo boxes ── */
QComboBox {{
    background: {C['card']};
    color: {C['text']};
    border: 1px solid {C['border']};
    border-radius: 7px;
    padding: 6px 10px;
    min-width: 140px;
}}
QComboBox:hover   {{ border-color: {C['text3']}; }}
QComboBox:focus   {{ border-color: {C['accent']}; }}
QComboBox:disabled {{ color: {C['text3']}; background: {C['surface']}; }}
QComboBox::drop-down {{ border: none; width: 28px; }}
QComboBox::down-arrow {{
    image: none;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 6px solid {C['text3']};
    margin-right: 8px;
}}
QComboBox QAbstractItemView {{
    background: {C['card']};
    color: {C['text']};
    border: 1px solid {C['border']};
    border-radius: 6px;
    selection-background-color: {C['accent']};
    selection-color: #ffffff;
    outline: none;
    padding: 4px;
}}
QAbstractItemView {{
    background: {C['card']};
    color: {C['text']};
    border: 1px solid {C['border']};
    border-radius: 6px;
    selection-background-color: {C['accent']};
    selection-color: #ffffff;
    outline: none;
    padding: 4px;
}}
QAbstractItemView::item {{
    padding: 4px 8px;
    color: {C['text']};
}}
QAbstractItemView::item:selected {{
    background: {C['accent']};
    color: #ffffff;
}}
QAbstractItemView::item:hover {{
    background: {C['card']};
    color: {C['text']};
}}

/* ── line edits ── */
QLineEdit {{
    background: {C['card']};
    color: {C['text']};
    border: 1px solid {C['border']};
    border-radius: 7px;
    padding: 6px 10px;
    min-width: 80px;
}}
QLineEdit:hover {{ border-color: {C['text3']}; }}
QLineEdit:focus {{ border-color: {C['accent']}; }}
QLineEdit:disabled {{ color: {C['text3']}; background: {C['surface']}; }}

/* ── console ── */
QTextEdit {{
    background: {C['console_bg']};
    color: {C['console_text']};
    border: 1px solid {C['border']};
    border-radius: 8px;
    padding: 10px;
    font-family: 'Cascadia Code', 'Consolas', 'Courier New', monospace;
    font-size: 12px;
}}

/* ── group boxes ── */
QGroupBox {{
    background: {C['surface']};
    border: 1px solid {C['border']};
    border-radius: 10px;
    margin-top: 14px;
    padding: 14px 12px 10px 12px;
    font-size: 11px;
    font-weight: 600;
    color: {C['text2']};
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 8px;
    left: 16px;
    color: {C['text2']};
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1.2px;
}}

/* ── progress bar ── */
QProgressBar {{
    background: {C['card']};
    border: none;
    border-radius: 5px;
    height: 8px;
    text-align: center;
    color: transparent;
}}
QProgressBar::chunk {{
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 {C['accent']}, stop:1 {C['success']});
    border-radius: 5px;
}}

/* ── status bar ── */
QStatusBar {{
    background: {C['header']};
    color: {C['text3']};
    border-top: 1px solid {C['border']};
    font-size: 12px;
}}

/* ── splitter ── */
QSplitter::handle {{ background: {C['border']}; width: 1px; height: 1px; }}

/* ── header bar (open file button area) ── */
#header_bar {{
    background: {C['header']};
    border-bottom: 1px solid {C['border']};
}}
"""


APP_STYLE = make_style(C)


# ── graph window ─────────────────────────────────────────────────────────────

class GraphWindow(QDialog):
    def __init__(self, fig, title: str, parent=None, style: str = ""):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(1200, 720)
        self.setStyleSheet(style or APP_STYLE)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        toolbar_frame = QFrame()
        toolbar_frame.setStyleSheet(f"background:{C['surface']}; border-bottom:1px solid {C['border']};")
        tlay = QHBoxLayout(toolbar_frame)
        tlay.setContentsMargins(8, 4, 8, 4)

        canvas = FigureCanvasQTAgg(fig)
        canvas.setStyleSheet("background: transparent;")

        toolbar = NavigationToolbar2QT(canvas, toolbar_frame)
        toolbar.setStyleSheet(f"""
            QToolBar {{ background: transparent; border: none; spacing: 4px; }}
            QToolButton {{
                background: {C['card']}; border: 1px solid {C['border']};
                border-radius: 6px; padding: 4px; color: {C['text']};
            }}
            QToolButton:hover {{ background: {C['border']}; }}
        """)
        tlay.addWidget(toolbar)
        tlay.addStretch()

        layout.addWidget(toolbar_frame)
        layout.addWidget(canvas, 1)
        canvas.draw()


# ── waveform player window ────────────────────────────────────────────────────

class WaveformPlayerWindow(QDialog):
    """Non-modal window: embedded waveform plot + audio playback with moving cursor."""

    def __init__(self, file_path: str, parent=None, style: str = ""):
        super().__init__(parent)
        self.setWindowTitle(f"Waveform  ·  {Path(file_path).name}")
        self.resize(1100, 620)
        self.setWindowModality(Qt.WindowModality.NonModal)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        self.setStyleSheet(style or APP_STYLE)

        self._file_path = file_path
        self._playing = False
        self._t_start = 0.0          # wall-clock time when play started
        self._play_offset = 0.0      # seconds into file at last play start
        self._duration = 0.0

        # ── load audio ──
        try:
            import soundfile as sf
            self._raw, self._fs = sf.read(file_path, always_2d=True)
        except Exception as e:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Load error", f"Cannot read file:\n{e}")
            self._raw = np.zeros((1, 1))
            self._fs = 44100

        n_ch = self._raw.shape[1]
        self._mono = self._raw.mean(axis=1)          # for cursor maths
        self._n_frames = len(self._mono)
        self._duration = self._n_frames / self._fs
        self._time_ax = np.linspace(0, self._duration, self._n_frames)

        # check sounddevice available
        self._sd_ok = False
        try:
            import sounddevice as _sd
            self._sd_ok = True
        except ImportError:
            pass

        self._build_ui(n_ch)

        # cursor timer (~30 fps)
        self._timer = QTimer(self)
        self._timer.setInterval(33)
        self._timer.timeout.connect(self._tick)

    # ── UI ──────────────────────────────────────────────────────────────────

    def _build_ui(self, n_ch: int):
        from matplotlib.figure import Figure
        from matplotlib.gridspec import GridSpec

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── matplotlib figure: waveform(s) + spectrogram ──
        n_wave = min(n_ch, 2)           # 1 or 2 waveform rows
        n_rows = n_wave + 1             # +1 for spectrogram
        fig_h = 2.2 * n_rows
        self._fig = Figure(figsize=(13, fig_h))
        self._fig.patch.set_facecolor(C['surface'])

        gs = GridSpec(n_rows, 1, figure=self._fig,
                      hspace=0.45, top=0.96, bottom=0.10,
                      left=0.07, right=0.97)

        self._axes = []
        self._cursors = []

        # ── waveform rows ──
        for i in range(n_wave):
            ax = self._fig.add_subplot(gs[i, 0])
            ax.set_facecolor(C['bg'])
            ch_data = self._raw[:, i]
            ax.plot(self._time_ax, ch_data,
                    color=C['accent2'], linewidth=0.6, alpha=0.85)
            peak = max(float(np.abs(ch_data).max()), 1e-6)
            ax.set_ylim(-peak * 1.05, peak * 1.05)
            ax.set_xlim(0, self._duration)
            title = "Waveform" if n_wave == 1 else f"Waveform – Ch {i + 1}"
            ax.set_title(title, color=C['text2'], fontsize=10, pad=3)
            ax.set_ylabel("Amplitude", color=C['text3'], fontsize=9)
            ax.tick_params(colors=C['text3'], labelsize=8)
            for sp in ax.spines.values():
                sp.set_edgecolor(C['border'])
            if i == n_wave - 1:
                ax.set_xlabel("")   # spectrogram below carries the time label
            # playback cursor
            cur, = ax.plot([0, 0], [-peak * 1.05, peak * 1.05],
                           color=C['danger'], linewidth=1.5, alpha=0.9)
            self._axes.append(ax)
            self._cursors.append(cur)

        # ── spectrogram row ──
        ax_spec = self._fig.add_subplot(gs[n_wave, 0])
        ax_spec.set_facecolor(C['bg'])

        # compute spectrogram on mono mix
        _, freqs, _, _ = ax_spec.specgram(
            self._mono, Fs=self._fs,
            NFFT=1024, noverlap=512,
            cmap='viridis', scale='dB'
        )
        ax_spec.set_yscale('log')
        if len(freqs) > 1:
            ax_spec.set_ylim(max(freqs[1], 20), freqs[-1])
        ax_spec.set_xlim(0, self._duration)
        ax_spec.set_title("Spectrogram (mono mix)", color=C['text2'], fontsize=10, pad=3)
        ax_spec.set_xlabel("Time [s]", color=C['text2'], fontsize=9)
        ax_spec.set_ylabel("Freq [Hz]", color=C['text3'], fontsize=9)
        ax_spec.tick_params(colors=C['text3'], labelsize=8)
        for sp in ax_spec.spines.values():
            sp.set_edgecolor(C['border'])

        # cursor on spectrogram
        ylim_spec = ax_spec.get_ylim()
        cur_spec, = ax_spec.plot([0, 0], [ylim_spec[0], ylim_spec[1]],
                                 color=C['danger'], linewidth=1.5, alpha=0.9)
        self._axes.append(ax_spec)
        self._cursors.append(cur_spec)
        self._ax_spec = ax_spec
        self._ylim_spec = ylim_spec

        self._canvas = FigureCanvasQTAgg(self._fig)
        self._canvas.setStyleSheet("background:transparent;")
        self._canvas.mpl_connect("button_press_event", self._on_canvas_click)

        toolbar_frame = QFrame()
        toolbar_frame.setStyleSheet(
            f"background:{C['surface']}; border-bottom:1px solid {C['border']};"
        )
        tlay = QHBoxLayout(toolbar_frame)
        tlay.setContentsMargins(8, 4, 8, 4)
        nav = NavigationToolbar2QT(self._canvas, toolbar_frame)
        nav.setStyleSheet(
            f"QToolBar {{ background:transparent; border:none; spacing:4px; }}"
            f"QToolButton {{ background:{C['card']}; border:1px solid {C['border']}; "
            f"border-radius:6px; padding:4px; color:{C['text']}; }}"
            f"QToolButton:hover {{ background:{C['border']}; }}"
        )
        tlay.addWidget(nav)
        tlay.addStretch()

        root.addWidget(toolbar_frame)
        root.addWidget(self._canvas, 1)
        root.addWidget(self._build_controls())
        self._canvas.draw()

    def _build_controls(self) -> QFrame:
        ctrl = QFrame()
        ctrl.setStyleSheet(
            f"background:{C['surface']}; border-top:1px solid {C['border']};"
        )
        lay = QHBoxLayout(ctrl)
        lay.setContentsMargins(16, 8, 16, 8)
        lay.setSpacing(10)

        # Play / Pause
        self._btn_play = QPushButton("▶   Play")
        self._btn_play.setFixedHeight(36)
        self._btn_play.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_play.clicked.connect(self._toggle_play)
        if not self._sd_ok:
            self._btn_play.setEnabled(False)
            self._btn_play.setToolTip("Install 'sounddevice' to enable playback")

        # Stop
        self._btn_stop = QPushButton("■   Stop")
        self._btn_stop.setFixedHeight(36)
        self._btn_stop.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_stop.clicked.connect(self._stop)
        if not self._sd_ok:
            self._btn_stop.setEnabled(False)

        # Time label
        self._lbl_time = QLabel(f"0.00  /  {self._duration:.2f} s")
        self._lbl_time.setStyleSheet(
            f"color:{C['text2']}; font-size:12px; background:transparent;"
        )

        # Hint
        hint = QLabel("  Click waveform to seek")
        hint.setStyleSheet(
            f"color:{C['text3']}; font-size:11px; font-style:italic; background:transparent;"
        )
        if not self._sd_ok:
            hint = QLabel("  ⚠  sounddevice not installed — pip install sounddevice")
            hint.setStyleSheet(
                f"color:{C['warning']}; font-size:11px; background:transparent;"
            )

        lay.addWidget(self._btn_play)
        lay.addWidget(self._btn_stop)
        lay.addSpacing(16)
        lay.addWidget(self._lbl_time)
        lay.addStretch()
        lay.addWidget(hint)
        return ctrl

    # ── playback ─────────────────────────────────────────────────────────────

    def _toggle_play(self):
        if self._playing:
            self._pause()
        else:
            self._play()

    def _play(self):
        import sounddevice as sd
        start_frame = max(0, min(int(self._play_offset * self._fs), self._n_frames - 1))
        if start_frame >= self._n_frames - 1:
            start_frame = 0
            self._play_offset = 0.0
        audio = self._raw[start_frame:]
        sd.play(audio, self._fs)
        self._t_start = time.perf_counter() - self._play_offset
        self._playing = True
        self._btn_play.setText("⏸   Pause")
        self._timer.start()

    def _pause(self):
        import sounddevice as sd
        self._play_offset = time.perf_counter() - self._t_start
        sd.stop()
        self._playing = False
        self._btn_play.setText("▶   Play")
        self._timer.stop()

    def _stop(self):
        import sounddevice as sd
        sd.stop()
        self._playing = False
        self._play_offset = 0.0
        self._btn_play.setText("▶   Play")
        self._timer.stop()
        self._set_cursor(0.0)
        self._lbl_time.setText(f"0.00  /  {self._duration:.2f} s")

    def _tick(self):
        elapsed = time.perf_counter() - self._t_start
        if elapsed >= self._duration:
            self._stop()
            return
        self._set_cursor(elapsed)
        self._lbl_time.setText(f"{elapsed:.2f}  /  {self._duration:.2f} s")

    def _set_cursor(self, t: float):
        for ax, cur in zip(self._axes, self._cursors):
            cur.set_xdata([t, t])
            # keep y-data spanning the current ylim (works for both linear and log axes)
            y0, y1 = ax.get_ylim()
            cur.set_ydata([y0, y1])
        self._canvas.draw_idle()

    def _on_canvas_click(self, event):
        if event.inaxes not in self._axes or event.xdata is None:
            return
        new_t = float(np.clip(event.xdata, 0, self._duration))
        was_playing = self._playing
        if was_playing:
            import sounddevice as sd
            sd.stop()
            self._timer.stop()
            self._playing = False
        self._play_offset = new_t
        self._set_cursor(new_t)
        self._lbl_time.setText(f"{new_t:.2f}  /  {self._duration:.2f} s")
        if was_playing:
            self._play()

    def closeEvent(self, event):
        self._timer.stop()
        if self._sd_ok:
            try:
                import sounddevice as sd
                sd.stop()
            except Exception:
                pass
        super().closeEvent(event)


# ── main application ─────────────────────────────────────────────────────────

class SQATModernApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("SQAT4PY  ·  Sound Quality Analysis Tool for Python")
        self.resize(1300, 880)
        self.setMinimumSize(960, 640)

        self._is_dark = True

        # state
        self.file_paths: List[str] = []
        self.file_label_map: Dict[str, str] = {}
        self.label_to_file: Dict[str, str] = {}
        self.current_file_index: int = 0
        self.save_dir: Optional[str] = None
        self.excel_dir: Optional[str] = None
        self.metric_params: Dict[str, Dict[str, Any]] = {
            m: dict(DEFAULTS.get(m, {})) for m in METRICS
        }
        self.current_config_metric: str = METRICS[0]
        self.metric_results: Dict[str, Dict[str, Any]] = {}
        self.last_metric: Optional[str] = None
        self.last_out: Optional[Dict[str, Any]] = None

        # threading
        self._worker: Optional[threading.Thread] = None
        self._result_q: queue.Queue = queue.Queue()
        self._analysis_finished = False
        self._total_steps = 1
        self._completed_steps = 0
        self._poll_timer = QTimer(self)
        self._poll_timer.setInterval(120)
        self._poll_timer.timeout.connect(self._poll_worker)

        QApplication.instance().setStyleSheet(APP_STYLE)
        self._build_ui()

    # ── UI construction ────────────────────────────────────────────────────

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        root_layout = QVBoxLayout(root)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        root_layout.addWidget(self._build_header())

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(1)
        splitter.addWidget(self._build_sidebar())
        splitter.addWidget(self._build_content())
        splitter.setSizes([270, 1030])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        root_layout.addWidget(splitter, 1)

        self._build_statusbar()
        self._sync_config_metric_choices(preferred=self.current_config_metric)
        self._refresh_params()
        self._apply_combo_popup_styles()

    def _build_header(self) -> QFrame:
        hdr = QFrame()
        hdr.setObjectName("header_bar")
        hdr.setFixedHeight(62)
        lay = QHBoxLayout(hdr)
        lay.setContentsMargins(20, 0, 20, 0)
        lay.setSpacing(12)

        # logo image
        self._logo_dark = str(Path(__file__).parent / "logos" / "logo_white.png")
        self._logo_light = str(Path(__file__).parent / "logos" / "logo.png")
        self.lbl_logo = QLabel()
        self.lbl_logo.setStyleSheet("background:transparent;")
        _px = QPixmap(self._logo_dark)
        if not _px.isNull():
            _px = _px.scaledToHeight(42, Qt.TransformationMode.SmoothTransformation)
            self.lbl_logo.setPixmap(_px)
            self.lbl_logo.setFixedSize(_px.width(), _px.height())
        lay.addWidget(self.lbl_logo)
        lay.addSpacing(24)

        # file chip
        self.lbl_file_chip = QLabel("  No files loaded  ")
        self.lbl_file_chip.setStyleSheet(
            f"background:rgba(99,102,241,0.12); color:{C['text3']}; "
            f"border-radius:10px; padding:4px 12px; font-size:11px; font-weight:600;"
        )

        # open button
        self.btn_pick = QPushButton("  Open WAV files…")
        self.btn_pick.setFixedHeight(36)
        self.btn_pick.setStyleSheet(
            f"QPushButton {{ background:{C['accent']}; color:#fff; border:none; "
            f"border-radius:8px; padding:0 18px; font-size:13px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{C['accent2']}; }}"
            f"QPushButton:pressed {{ background:#4338ca; }}"
        )
        self.btn_pick.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_pick.clicked.connect(self.on_pick_files)

        # active file
        self._lbl_af = QLabel("Active file:")
        self._lbl_af.setStyleSheet(f"background:transparent; color:{C['text3']}; font-size:12px;")
        self.cmb_active_file = QComboBox()
        self.cmb_active_file.setFixedWidth(210)
        self.cmb_active_file.setEnabled(False)
        self.cmb_active_file.currentIndexChanged.connect(self._on_active_file_changed)

        # plot metric
        self._lbl_pm = QLabel("Plot metric:")
        self._lbl_pm.setStyleSheet(f"background:transparent; color:{C['text3']}; font-size:12px;")
        self.cmb_plot_metric = QComboBox()
        self.cmb_plot_metric.setFixedWidth(210)
        self.cmb_plot_metric.setEnabled(False)
        self.cmb_plot_metric.currentIndexChanged.connect(lambda _: self._refresh_plot_target())

        lay.addWidget(self.lbl_file_chip)
        lay.addWidget(self.btn_pick)
        lay.addSpacing(16)

        self._hdr_sep = QFrame()
        self._hdr_sep.setFrameShape(QFrame.Shape.VLine)
        self._hdr_sep.setStyleSheet(f"color:{C['border']};")
        lay.addWidget(self._hdr_sep)
        lay.addSpacing(8)

        lay.addWidget(self._lbl_af)
        lay.addWidget(self.cmb_active_file)
        lay.addSpacing(16)
        lay.addWidget(self._lbl_pm)
        lay.addWidget(self.cmb_plot_metric)
        lay.addStretch()

        self.btn_theme = QPushButton("☀  Light")
        self.btn_theme.setFixedHeight(32)
        self.btn_theme.setFixedWidth(90)
        self.btn_theme.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_theme.setStyleSheet(
            f"QPushButton {{ background:transparent; color:{C['text3']}; border:1px solid {C['border']}; "
            f"border-radius:16px; font-size:12px; font-weight:500; padding:0 10px; }}"
            f"QPushButton:hover {{ color:{C['text']}; border-color:{C['text3']}; }}"
        )
        self.btn_theme.clicked.connect(self._toggle_theme)
        lay.addWidget(self.btn_theme)

        return hdr

    def _build_sidebar(self) -> QFrame:
        self._sidebar_frame = QFrame()
        sb = self._sidebar_frame
        sb.setObjectName("sidebar")
        sb.setFixedWidth(272)
        sb.setStyleSheet(f"background:{C['surface']}; border-right:1px solid {C['border']};")

        lay = QVBoxLayout(sb)
        lay.setContentsMargins(0, 18, 0, 18)
        lay.setSpacing(0)

        # section header
        self._lbl_metrics_header = QLabel("METRICS TO ANALYZE")
        self._lbl_metrics_header.setStyleSheet(
            f"background:transparent; color:{C['text3']}; font-size:10px; "
            f"font-weight:700; letter-spacing:1.8px; padding:0 16px 8px 16px;"
        )
        lay.addWidget(self._lbl_metrics_header)

        # helper tip
        self._lbl_metrics_tip = QLabel("Ctrl+click for multi-select")
        self._lbl_metrics_tip.setStyleSheet(
            f"background:transparent; color:{C['text3']}; font-size:10px; "
            f"font-style:italic; padding:0 16px 10px 16px;"
        )
        lay.addWidget(self._lbl_metrics_tip)

        # metrics list
        self.lst_metrics = QListWidget()
        self.lst_metrics.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.lst_metrics.setSpacing(0)
        self.lst_metrics.setFont(QFont("Segoe UI", 12))
        self.lst_metrics.setStyleSheet(
            f"QListWidget {{ background:{C['surface']}; border:none; padding:0 6px; }}"
            f"QListWidget::item {{ padding:10px 10px; border-radius:6px; margin:1px 2px; "
            f"color:{C['text2']}; border-left:3px solid transparent; }}"
            f"QListWidget::item:hover {{ background:{C['card']}; color:{C['text']}; }}"
            f"QListWidget::item:selected {{ background:rgba(99,102,241,0.18); "
            f"color:{C['accent2']}; border-left:3px solid {C['accent']}; }}"
        )
        for m in METRICS:
            self.lst_metrics.addItem(QListWidgetItem(m))
        self.lst_metrics.setCurrentRow(0)
        self.lst_metrics.itemSelectionChanged.connect(self._on_metrics_selection_changed)
        lay.addWidget(self.lst_metrics, 1)

        lay.addSpacing(14)
        self._sidebar_sep = QFrame()
        self._sidebar_sep.setFrameShape(QFrame.Shape.HLine)
        self._sidebar_sep.setStyleSheet(f"color:{C['border']}; background:{C['border']};")
        self._sidebar_sep.setFixedHeight(1)
        lay.addWidget(self._sidebar_sep)
        lay.addSpacing(14)

        # actions section
        self._lbl_actions_header = QLabel("ACTIONS")
        self._lbl_actions_header.setStyleSheet(
            f"background:transparent; color:{C['text3']}; font-size:10px; "
            f"font-weight:700; letter-spacing:1.8px; padding:0 16px 10px 16px;"
        )
        lay.addWidget(self._lbl_actions_header)

        btn_area = QWidget()
        btn_area.setStyleSheet("background:transparent;")
        blay = QVBoxLayout(btn_area)
        blay.setContentsMargins(14, 0, 18, 0)
        blay.setSpacing(8)

        self.btn_analyze = QPushButton("▶   Run Analysis")
        self.btn_analyze.setObjectName("btn_run")
        self.btn_analyze.setFixedHeight(48)
        self.btn_analyze.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_analyze.clicked.connect(self.on_analyze)
        self.btn_analyze.setStyleSheet(self._style_btn_run())
        blay.addWidget(self.btn_analyze)

        self.btn_plot = QPushButton("⬛   Open Graphs Window")
        self.btn_plot.setObjectName("btn_graphs")
        self.btn_plot.setFixedHeight(38)
        self.btn_plot.setEnabled(False)
        self.btn_plot.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_plot.clicked.connect(self.open_graph_window)
        self.btn_plot.setStyleSheet(self._style_btn_graphs())
        blay.addWidget(self.btn_plot)

        self.btn_wave = QPushButton("〰   Waveform & Play")
        self.btn_wave.setObjectName("btn_wave")
        self.btn_wave.setFixedHeight(38)
        self.btn_wave.setEnabled(False)
        self.btn_wave.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_wave.clicked.connect(self.on_view_waveform)
        self.btn_wave.setStyleSheet(self._style_btn_wave())
        blay.addWidget(self.btn_wave)

        lay.addWidget(btn_area)
        return sb

    def _build_content(self) -> QWidget:
        panel = QWidget()
        self._content_panel = panel
        # no inline setStyleSheet — QWidget rule in APP_STYLE covers the background
        # and avoids creating a CSS-scope boundary that blocks popup styling
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(20, 14, 20, 14)
        lay.setSpacing(10)

        # ── options ──
        opts = QGroupBox("OPTIONS")
        self._opts_group = opts
        opts.setMaximumHeight(72)
        olay = QHBoxLayout(opts)
        olay.setContentsMargins(16, 6, 16, 6)
        olay.setSpacing(24)
        self.chk_show  = QCheckBox("Show plots after run")
        self.chk_save  = QCheckBox("Save figures")
        self.chk_excel = QCheckBox("Export to Excel")
        self.chk_split = QCheckBox("Split figures")
        for chk in (self.chk_show, self.chk_save, self.chk_excel, self.chk_split):
            olay.addWidget(chk)
        olay.addStretch()
        self.chk_save.stateChanged.connect(self.on_save_toggle)
        self.chk_excel.stateChanged.connect(self.on_excel_toggle)
        lay.addWidget(opts)

        # ── parameters ──
        self.params_group = QGroupBox("PARAMETERS")
        self.params_group.setMaximumHeight(150)
        params_outer = QVBoxLayout(self.params_group)
        params_outer.setContentsMargins(16, 8, 16, 8)
        params_outer.setSpacing(8)

        edit_row = QHBoxLayout()
        edit_row.setSpacing(10)
        self._lbl_edit_params = QLabel("Edit parameters for:")
        self._lbl_edit_params.setStyleSheet(f"color:{C['text3']}; font-size:12px; background:transparent;")
        self.cmb_config_metric = QComboBox()
        self.cmb_config_metric.setFixedWidth(290)
        self.cmb_config_metric.currentIndexChanged.connect(self._on_config_metric_changed)
        edit_row.addWidget(self._lbl_edit_params)
        edit_row.addWidget(self.cmb_config_metric)
        edit_row.addStretch()
        params_outer.addLayout(edit_row)

        self.params_widget = QWidget()
        self.params_widget.setStyleSheet("background:transparent;")
        self.params_grid = QGridLayout(self.params_widget)
        self.params_grid.setContentsMargins(0, 2, 0, 0)
        self.params_grid.setHorizontalSpacing(16)
        self.params_grid.setVerticalSpacing(6)
        params_outer.addWidget(self.params_widget)
        self._build_param_widgets()
        lay.addWidget(self.params_group)

        # ── console ──
        console_group = QGroupBox("CONSOLE OUTPUT")
        self._console_group = console_group
        clay = QVBoxLayout(console_group)
        clay.setContentsMargins(8, 8, 8, 8)
        clay.setSpacing(0)
        self.txt = QTextEdit()
        self.txt.setReadOnly(True)
        self.txt.setFont(QFont("Cascadia Code", 11))
        self.txt.setPlaceholderText(
            "Select WAV files, choose metrics, configure parameters, then click Run Analysis."
        )
        clay.addWidget(self.txt)
        lay.addWidget(console_group, 1)

        return panel

    def _build_statusbar(self):
        sb = QStatusBar()
        self._statusbar = sb
        sb.setStyleSheet(
            f"QStatusBar {{ background:{C['header']}; color:{C['text3']}; "
            f"border-top:1px solid {C['border']}; font-size:12px; }}"
        )
        self.setStatusBar(sb)

        self.progress = QProgressBar()
        self.progress.setFixedHeight(8)
        self.progress.setFixedWidth(340)
        self.progress.setTextVisible(False)
        self.progress.setValue(0)

        self.lbl_status = QLabel("Ready")
        self.lbl_status.setStyleSheet(f"color:{C['text3']}; font-size:12px; padding:0 16px;")

        sb.addWidget(QLabel("  "))
        sb.addWidget(self.progress)
        sb.addPermanentWidget(self.lbl_status)

    def _combo_popup_css(self) -> str:
        return (
            f"background:{C['card']}; color:{C['text']}; "
            f"border:1px solid {C['border']}; border-radius:6px; "
            f"selection-background-color:{C['accent']}; selection-color:#ffffff; "
            f"outline:none; padding:4px;"
        )

    def _apply_combo_popup_styles(self):
        """Set popup view stylesheet directly on every combobox — bypasses all cascade issues."""
        css = self._combo_popup_css()
        for combo in [self.cmb_active_file, self.cmb_plot_metric, self.cmb_config_metric]:
            combo.view().setStyleSheet(css)
        for ctrl in getattr(self, "param_controls", {}).values():
            if ctrl["kind"] == "combo":
                ctrl["widget"].view().setStyleSheet(css)

    def _style_btn_run(self) -> str:
        return (
            f"QPushButton {{ background:{C['success']}; color:#ffffff; "
            f"border:2px solid {C['success2']}; border-radius:10px; "
            f"padding:12px 24px; font-size:15px; font-weight:700; "
            f"letter-spacing:0.5px; margin-right:4px; }}"
            f"QPushButton:hover {{ background:{C['success2']}; border-color:{C['success']}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:#059669; border-color:#047857; color:#ffffff; }}"
            f"QPushButton:disabled {{ background:transparent; color:{C['text3']}; "
            f"border:2px solid {C['border']}; margin-right:4px; }}"
        )

    def _style_btn_graphs(self) -> str:
        return (
            f"QPushButton {{ background:rgba(99,102,241,0.15); color:{C['accent2']}; "
            f"border:1px solid rgba(99,102,241,0.5); border-radius:8px; "
            f"padding:9px 18px; font-size:13px; font-weight:600; margin-right:4px; }}"
            f"QPushButton:hover {{ background:rgba(99,102,241,0.28); border-color:{C['accent']}; color:{C['accent2']}; }}"
            f"QPushButton:disabled {{ background:transparent; color:{C['text3']}; "
            f"border:1px solid {C['border']}; margin-right:4px; }}"
        )

    def _style_btn_wave(self) -> str:
        return (
            f"QPushButton {{ background:rgba(245,158,11,0.12); color:{C['warning']}; "
            f"border:1px solid rgba(245,158,11,0.5); border-radius:8px; "
            f"padding:9px 18px; font-size:13px; font-weight:600; margin-right:4px; }}"
            f"QPushButton:hover {{ background:rgba(245,158,11,0.22); border-color:{C['warning']}; color:{C['warning']}; }}"
            f"QPushButton:disabled {{ background:transparent; color:{C['text3']}; "
            f"border:1px solid {C['border']}; margin-right:4px; }}"
        )

    def _ensure_imports(self):
        """Synchronously load heavy modules if not yet done (used by waveform button)."""
        global _imports_done
        if not _imports_done:
            _lazy_imports()
            _imports_done = True

    def _toggle_theme(self):
        self._is_dark = not self._is_dark
        self._apply_theme()

    def _apply_theme(self):
        global C, APP_STYLE
        C = dict(DARK_COLORS if self._is_dark else LIGHT_COLORS)
        APP_STYLE = make_style(C)
        QApplication.instance().setStyleSheet(APP_STYLE)

        # ── logo ──
        logo_path = self._logo_dark if self._is_dark else self._logo_light
        _px = QPixmap(logo_path)
        if not _px.isNull():
            _px = _px.scaledToHeight(42, Qt.TransformationMode.SmoothTransformation)
            self.lbl_logo.setPixmap(_px)
            self.lbl_logo.setFixedSize(_px.width(), _px.height())

        # ── toggle button ──
        self.btn_theme.setText("☀  Light" if self._is_dark else "🌙  Dark")
        self.btn_theme.setStyleSheet(
            f"QPushButton {{ background:transparent; color:{C['text3']}; border:1px solid {C['border']}; "
            f"border-radius:16px; font-size:12px; font-weight:500; padding:0 10px; }}"
            f"QPushButton:hover {{ color:{C['text']}; border-color:{C['text3']}; }}"
        )

        # ── header inline widgets ──
        self._hdr_sep.setStyleSheet(f"color:{C['border']};")
        self._lbl_af.setStyleSheet(f"background:transparent; color:{C['text3']}; font-size:12px;")
        self._lbl_pm.setStyleSheet(f"background:transparent; color:{C['text3']}; font-size:12px;")
        self.btn_pick.setStyleSheet(
            f"QPushButton {{ background:{C['accent']}; color:#fff; border:none; "
            f"border-radius:8px; padding:0 18px; font-size:13px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{C['accent2']}; }}"
            f"QPushButton:pressed {{ background:#4338ca; }}"
        )
        if self.file_paths:
            self.lbl_file_chip.setStyleSheet(
                f"background:rgba(16,185,129,0.18); color:{C['success2']}; "
                f"border-radius:10px; padding:4px 12px; font-size:11px; font-weight:600;"
            )
        else:
            self.lbl_file_chip.setStyleSheet(
                f"background:rgba(99,102,241,0.12); color:{C['text3']}; "
                f"border-radius:10px; padding:4px 12px; font-size:11px; font-weight:600;"
            )

        # ── sidebar inline widgets ──
        self._sidebar_frame.setStyleSheet(
            f"background:{C['surface']}; border-right:1px solid {C['border']};"
        )
        self._lbl_metrics_header.setStyleSheet(
            f"background:transparent; color:{C['text3']}; font-size:10px; "
            f"font-weight:700; letter-spacing:1.8px; padding:0 16px 8px 16px;"
        )
        self._lbl_metrics_tip.setStyleSheet(
            f"background:transparent; color:{C['text3']}; font-size:10px; "
            f"font-style:italic; padding:0 16px 10px 16px;"
        )
        self.lst_metrics.setStyleSheet(
            f"QListWidget {{ background:{C['surface']}; border:none; padding:0 6px; }}"
            f"QListWidget::item {{ padding:10px 10px; border-radius:6px; margin:1px 2px; "
            f"color:{C['text2']}; border-left:3px solid transparent; }}"
            f"QListWidget::item:hover {{ background:{C['card']}; color:{C['text']}; }}"
            f"QListWidget::item:selected {{ background:rgba(99,102,241,0.18); "
            f"color:{C['accent2']}; border-left:3px solid {C['accent']}; }}"
        )
        self._sidebar_sep.setStyleSheet(f"color:{C['border']}; background:{C['border']};")
        self._lbl_actions_header.setStyleSheet(
            f"background:transparent; color:{C['text3']}; font-size:10px; "
            f"font-weight:700; letter-spacing:1.8px; padding:0 16px 10px 16px;"
        )

        # ── content panel ──
        self._lbl_edit_params.setStyleSheet(
            f"color:{C['text3']}; font-size:12px; background:transparent;"
        )

        # ── status bar ──
        self._statusbar.setStyleSheet(
            f"QStatusBar {{ background:{C['header']}; color:{C['text3']}; "
            f"border-top:1px solid {C['border']}; font-size:12px; }}"
        )
        self.lbl_status.setStyleSheet(f"color:{C['text3']}; font-size:12px; padding:0 16px;")

        # ── param label colors ──
        for ctrl in getattr(self, "param_controls", {}).values():
            ctrl["label"].setStyleSheet(f"color:{C['text2']}; background:transparent; font-size:12px;")

        # ── action buttons (direct inline style guarantees correct colours) ──
        self.btn_analyze.setStyleSheet(self._style_btn_run())
        self.btn_plot.setStyleSheet(self._style_btn_graphs())
        self.btn_wave.setStyleSheet(self._style_btn_wave())

        # ── combobox popups (direct view style bypasses CSS-scope boundaries) ──
        self._apply_combo_popup_styles()

        # ── console ──
        self.txt.setStyleSheet(
            f"background:{C['console_bg']}; color:{C['console_text']}; "
            f"border:1px solid {C['border']}; border-radius:8px; padding:10px; "
            f"font-family:'Cascadia Code','Consolas','Courier New',monospace; font-size:12px;"
        )

    def _build_param_widgets(self):
        self.param_controls: Dict[str, Dict[str, Any]] = {}
        for key, spec in PARAM_SPECS.items():
            lbl = QLabel(spec["label"] + ":")
            lbl.setStyleSheet(f"color:{C['text2']}; background:transparent; font-size:12px;")
            lbl.hide()
            if spec["kind"] == "entry":
                widget = QLineEdit()
                widget.setFixedWidth(110)
                widget.hide()
                self.param_controls[key] = {"label": lbl, "widget": widget, "spec": spec, "kind": "entry"}
            else:
                widget = QComboBox()
                widget.addItems([d for d, _ in spec["options"]])
                widget.setFixedWidth(200)
                widget.hide()
                self.param_controls[key] = {"label": lbl, "widget": widget, "spec": spec, "kind": "combo"}

    # ── parameter helpers ──────────────────────────────────────────────────

    def _get_selected_metrics(self) -> List[str]:
        return [self.lst_metrics.item(i).text()
                for i in range(self.lst_metrics.count())
                if self.lst_metrics.item(i).isSelected()]

    def _sync_config_metric_choices(self, preferred: Optional[str] = None):
        selected = self._get_selected_metrics()
        editable = selected or [self.current_config_metric or METRICS[0]]
        self.cmb_config_metric.blockSignals(True)
        self.cmb_config_metric.clear()
        self.cmb_config_metric.addItems(editable)
        target = preferred if preferred in editable else editable[0]
        self.current_config_metric = target
        idx = self.cmb_config_metric.findText(target)
        self.cmb_config_metric.setCurrentIndex(max(idx, 0))
        self.cmb_config_metric.blockSignals(False)
        self._refresh_params()

    def _on_metrics_selection_changed(self):
        self._store_current_param_state()
        selected = self._get_selected_metrics()
        preferred = (
            self.current_config_metric if self.current_config_metric in selected
            else (selected[0] if selected else self.current_config_metric)
        )
        self._sync_config_metric_choices(preferred=preferred)

    def _on_config_metric_changed(self):
        prev = getattr(self, "current_config_metric", None)
        self._store_current_param_state(prev)
        picked = self.cmb_config_metric.currentText().strip()
        if picked:
            self.current_config_metric = picked
        self._refresh_params()

    def _control_value_to_storage(self, key: str):
        ctrl = self.param_controls[key]
        if ctrl["kind"] == "entry":
            return ctrl["widget"].text().strip()
        display = ctrl["widget"].currentText()
        for disp, val in ctrl["spec"]["options"]:
            if display == disp:
                return val
        return display

    def _set_control_value(self, key: str, value: Any):
        ctrl = self.param_controls[key]
        spec = ctrl["spec"]
        if ctrl["kind"] == "entry":
            ctrl["widget"].setText("" if value is None else str(value))
            return
        reverse = {stored: disp for disp, stored in spec["options"]}
        display = reverse.get(value)
        if display is None and spec["options"]:
            display = spec["options"][0][0]
        idx = ctrl["widget"].findText(display)
        if idx >= 0:
            ctrl["widget"].setCurrentIndex(idx)

    def _store_current_param_state(self, metric: Optional[str] = None):
        metric = metric or getattr(self, "current_config_metric", None)
        if not metric:
            return
        params = dict(self.metric_params.get(metric, {}))
        for key in METRIC_PARAM_LAYOUT.get(metric, []):
            params[key] = self._control_value_to_storage(key)
        self.metric_params[metric] = params

    def _refresh_params(self):
        metric = getattr(self, "current_config_metric", METRICS[0])
        defaults = dict(DEFAULTS.get(metric, {}))
        stored   = dict(self.metric_params.get(metric, {}))
        merged   = {**defaults, **stored}
        self.metric_params[metric] = merged

        for ctrl in self.param_controls.values():
            ctrl["label"].hide()
            ctrl["widget"].hide()
            self.params_grid.removeWidget(ctrl["label"])
            self.params_grid.removeWidget(ctrl["widget"])

        for col, key in enumerate(METRIC_PARAM_LAYOUT.get(metric, [])):
            ctrl = self.param_controls[key]
            self.params_grid.addWidget(ctrl["label"],  0, col)
            self.params_grid.addWidget(ctrl["widget"], 1, col)
            ctrl["label"].show()
            ctrl["widget"].show()
            self._set_control_value(key, merged.get(key, ""))

        self.params_group.setTitle(f"PARAMETERS  ·  {metric}")

    # ── file helpers ───────────────────────────────────────────────────────

    def _build_file_label_maps(self):
        basenames = [Path(p).name for p in self.file_paths]
        dupes = {n for n in basenames if basenames.count(n) > 1}
        self.file_label_map = {}
        self.label_to_file  = {}
        for path in self.file_paths:
            label = str(path) if Path(path).name in dupes else Path(path).name
            self.file_label_map[path] = label
            self.label_to_file[label] = path

    def _selected_active_file_path(self) -> Optional[str]:
        label = self.cmb_active_file.currentText().strip()
        if label and label in self.label_to_file:
            return self.label_to_file[label]
        if self.file_paths:
            return self.file_paths[min(self.current_file_index, len(self.file_paths) - 1)]
        return None

    def _on_active_file_changed(self):
        fp = self._selected_active_file_path()
        if fp and fp in self.file_paths:
            self.current_file_index = self.file_paths.index(fp)

    def _append_terminal(self, text: str, hex_color: Optional[str] = None):
        cursor = self.txt.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        self.txt.setTextCursor(cursor)
        if hex_color:
            self.txt.setTextColor(QColor(hex_color))
        self.txt.insertPlainText(text.rstrip() + "\n")
        if hex_color:
            self.txt.setTextColor(QColor("#7dd3a8"))
        self.txt.ensureCursorVisible()

    # ── event handlers ─────────────────────────────────────────────────────

    def on_pick_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Choose WAV files", "", "WAV files (*.wav);;All files (*.*)"
        )
        if not paths:
            return
        self.file_paths = paths
        self.current_file_index = 0
        self._build_file_label_maps()
        n = len(paths)
        self.lbl_file_chip.setText(f"  {n} file{'s' if n != 1 else ''} loaded  ")
        self.lbl_file_chip.setStyleSheet(
            f"background:rgba(16,185,129,0.18); color:{C['success2']}; "
            f"border-radius:10px; padding:4px 12px; font-size:11px; font-weight:600;"
        )
        labels = [self.file_label_map[p] for p in self.file_paths]
        self.cmb_active_file.blockSignals(True)
        self.cmb_active_file.clear()
        self.cmb_active_file.addItems(labels)
        self.cmb_active_file.setCurrentIndex(0)
        self.cmb_active_file.blockSignals(False)
        self.cmb_active_file.setEnabled(True)
        self.btn_wave.setEnabled(True)
        self.lbl_status.setText(f"Loaded {n} file(s)")

    def on_save_toggle(self, state: int):
        if state == Qt.CheckState.Checked.value or state == 2:
            path = QFileDialog.getExistingDirectory(self, "Choose directory to save figures")
            if not path:
                self.chk_save.blockSignals(True)
                self.chk_save.setChecked(False)
                self.chk_save.blockSignals(False)
                return
            self.save_dir = path
        else:
            self.save_dir = None

    def on_excel_toggle(self, state: int):
        if state == Qt.CheckState.Checked.value or state == 2:
            path = QFileDialog.getExistingDirectory(self, "Choose directory to save Excel file")
            if not path:
                self.chk_excel.blockSignals(True)
                self.chk_excel.setChecked(False)
                self.chk_excel.blockSignals(False)
                return
            self.excel_dir = path
        else:
            self.excel_dir = None

    def on_view_waveform(self):
        if not self.file_paths:
            return
        self._ensure_imports()
        fp = self._selected_active_file_path() or self.file_paths[0]
        try:
            dlg = WaveformPlayerWindow(fp, parent=self, style=APP_STYLE)
            dlg.show()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to show waveform:\n{e}")

    def _set_running(self, running: bool):
        ok = not running
        self.btn_analyze.setEnabled(ok)
        self.btn_pick.setEnabled(ok)
        self.lst_metrics.setEnabled(ok)
        for chk in (self.chk_show, self.chk_save, self.chk_excel, self.chk_split):
            chk.setEnabled(ok)
        self.cmb_active_file.setEnabled(ok and bool(self.file_paths))
        self.cmb_config_metric.setEnabled(ok)
        self.cmb_plot_metric.setEnabled(ok and bool(self.metric_results))
        self.btn_wave.setEnabled(ok and bool(self.file_paths))
        self.btn_plot.setEnabled(ok and bool(self.metric_results))

        if running:
            self.progress.setMaximum(max(self._total_steps, 1))
            self.progress.setValue(0)
            self.lbl_status.setText(f"0 / {self._total_steps}")
            self.setCursor(Qt.CursorShape.WaitCursor)
        else:
            self.unsetCursor()
            if self._analysis_finished:
                self.lbl_status.setText("Done  ✓")
            elif self._completed_steps:
                self.lbl_status.setText(f"{self._completed_steps} / {self._total_steps}")
            else:
                self.lbl_status.setText("Ready")

    # ── analysis ───────────────────────────────────────────────────────────

    def _resolve_metric_params(self, metric: str) -> Dict[str, Any]:
        raw = {**DEFAULTS.get(metric, {}), **self.metric_params.get(metric, {})}
        p: Dict[str, Any] = {}

        def pf(name: str, default: float = 0.0) -> float:
            v = raw.get(name, default)
            return float(default) if v in (None, "") else float(v)

        layout = METRIC_PARAM_LAYOUT.get(metric, [])
        if "time_skip"   in layout: p["time_skip"]   = pf("time_skip", DEFAULTS.get(metric,{}).get("time_skip", 0.0))
        if "field"       in layout: p["field"]       = int(raw.get("field", DEFAULTS.get(metric,{}).get("field", 0)))
        if "method"      in layout: p["method"]      = int(raw.get("method", DEFAULTS.get(metric,{}).get("method", 1)))
        if "weight_type" in layout: p["weight_type"] = str(raw.get("weight_type", "DIN45692"))
        if "method_fs"   in layout: p["method_fs"]   = int(raw.get("method_fs", DEFAULTS.get(metric,{}).get("method_fs", 1)))
        if metric == "EPNL (FAR Part 36)":
            p["method_epnl"] = 1
            tr = raw.get("threshold_epnl", None)
            p["threshold_epnl"] = None if tr in (None, "") else float(tr)
        return p

    def on_analyze(self):
        if not self.file_paths:
            QMessageBox.information(self, "Select files", "Please choose WAV files first.")
            return
        self._store_current_param_state()
        selected_metrics = self._get_selected_metrics()
        if not selected_metrics:
            QMessageBox.information(self, "Select metrics", "Please choose at least one metric.")
            return
        try:
            params_by_metric = {m: self._resolve_metric_params(m) for m in selected_metrics}
        except Exception as e:
            QMessageBox.critical(self, "Invalid parameters", f"Review metric parameters:\n{e}")
            return

        want_plots = self.chk_show.isChecked()
        want_save  = self.chk_save.isChecked()
        want_excel = self.chk_excel.isChecked()
        want_split = self.chk_split.isChecked()

        self.metric_results = {}
        self.last_metric = None
        self.last_out = None
        self._result_q = queue.Queue()
        self._analysis_finished = False
        self._completed_steps = 0
        self._total_steps = max(len(self.file_paths) * len(selected_metrics), 1)

        self._set_running(True)
        self.txt.clear()
        self._append_terminal(f"Files    : {len(self.file_paths)}", C["text2"])
        self._append_terminal(f"Metrics  : {', '.join(selected_metrics)}", C["text2"])
        self._append_terminal("─" * 60, C["text3"])

        self._worker = threading.Thread(
            target=self._analyze_worker,
            args=(selected_metrics, params_by_metric, want_plots, want_save, want_excel, want_split),
            daemon=True,
        )
        self._worker.start()
        self._poll_timer.start()

    # ── threading ─────────────────────────────────────────────────────────

    def _analyze_worker(self, selected_metrics, params_by_metric, want_plots, want_save, want_excel, want_split):
        try:
            global _imports_done
            if not _imports_done:
                self._result_q.put(("log", "Loading modules…", C["text3"]))
                _lazy_imports()
                _imports_done = True
            results_by_metric = {m: [] for m in selected_metrics}
            total = len(self.file_paths) * len(selected_metrics)
            done = 0
            for fi, fp in enumerate(self.file_paths, 1):
                fname = Path(fp).name
                self._result_q.put(("log", f"[{fi}/{len(self.file_paths)}]  {fname}", None))
                insig = fs = None
                load_err = None
                try:
                    insig, fs = wav2sig(fp)
                except Exception as e:
                    load_err = str(e)
                    self._result_q.put(("log", f"    ✗ Load failed: {load_err}", C["danger"]))

                for mi, metric in enumerate(selected_metrics, 1):
                    self._result_q.put(("log", f"    ({mi}/{len(selected_metrics)}) {metric}…", C["text3"]))
                    try:
                        if load_err:
                            raise RuntimeError(load_err)
                        out = self._run_metric_worker(metric, params_by_metric[metric], insig, fs, show_plots=False)
                        results_by_metric[metric].append((fp, out, None))
                        summary = self._summarize(metric, out).splitlines()[0]
                        self._result_q.put(("log", f"        ✓ {summary}", C["success2"]))
                    except Exception as e:
                        results_by_metric[metric].append((fp, None, str(e)))
                        self._result_q.put(("log", f"        ✗ {e}", C["danger"]))
                    done += 1
                    self._result_q.put(("progress", done, total))

            self._result_q.put(("ok", results_by_metric, want_plots, want_save, want_excel, want_split))
        except Exception as e:
            self._result_q.put(("err", str(e)))

    def _poll_worker(self):
        finished = False
        while True:
            try:
                msg = self._result_q.get_nowait()
            except queue.Empty:
                break

            kind = msg[0]
            if kind == "log":
                self._append_terminal(msg[1], msg[2] if len(msg) > 2 else None)
            elif kind == "progress":
                _, step, total = msg
                self._completed_steps = step
                self._total_steps = max(total, 1)
                self.progress.setMaximum(self._total_steps)
                self.progress.setValue(step)
                self.lbl_status.setText(f"{step} / {self._total_steps}")
            elif kind == "ok":
                finished = True
                _, results_by_metric, wp, ws, we, wsp = msg
                self._handle_success(results_by_metric, wp, ws, we, wsp)
            elif kind == "err":
                finished = True
                self._analysis_finished = False
                self._poll_timer.stop()
                self._set_running(False)
                QMessageBox.critical(self, "Error", f"Analysis failed:\n{msg[1]}")

        if finished:
            self._poll_timer.stop()
            return

        if self._worker and not self._worker.is_alive():
            if not self._analysis_finished:
                self._poll_timer.stop()
                self._set_running(False)

    def _handle_success(self, results_by_metric, want_plots, want_save, want_excel, want_split):
        self.metric_results = {}
        self._append_terminal("\n" + "─" * 60, C["text3"])
        self._append_terminal("  SUMMARY", C["accent2"])
        self._append_terminal("─" * 60, C["text3"])

        last_ok_metric = last_ok_out = None
        for metric, results in results_by_metric.items():
            self._append_terminal(f"\n  {metric}", C["warning"])
            self.metric_results[metric] = {}
            for fp, out, err in results:
                fname = self.file_label_map.get(fp, Path(fp).name)
                if err:
                    self._append_terminal(f"    ✗ {fname}: {err}", C["danger"])
                    continue
                self.metric_results[metric][fp] = out
                summary = self._summarize(metric, out).replace("\n", "  |  ")
                self._append_terminal(f"    ✓ {fname}:  {summary}", C["success2"])
                if want_save and self.save_dir:
                    self._save_figure(metric, out, fp, want_split)
                last_ok_metric = metric
                last_ok_out = out
            if not self.metric_results[metric]:
                del self.metric_results[metric]

        if want_excel and self.excel_dir:
            self._export_results_excel(results_by_metric)

        self.last_metric = last_ok_metric
        self.last_out = last_ok_out
        self._update_plot_metric_choices()
        self._analysis_finished = True
        self._set_running(False)

        if want_plots and self.metric_results:
            self.open_graph_window()
        if want_save and self.save_dir:
            self._append_terminal(f"\n  Figures saved → {self.save_dir}", C["success2"])

    def _update_plot_metric_choices(self):
        available = list(self.metric_results.keys())
        self.cmb_plot_metric.blockSignals(True)
        self.cmb_plot_metric.clear()
        if available:
            self.cmb_plot_metric.addItems(available)
            self.cmb_plot_metric.setEnabled(True)
            self.btn_plot.setEnabled(True)
            self._refresh_plot_target()
        else:
            self.cmb_plot_metric.setEnabled(False)
            self.btn_plot.setEnabled(False)
        self.cmb_plot_metric.blockSignals(False)

    def _refresh_plot_target(self):
        metric = self.cmb_plot_metric.currentText().strip()
        if not metric or metric not in self.metric_results:
            return
        available_files = list(self.metric_results[metric].keys())
        if not available_files:
            return
        current_path = self._selected_active_file_path()
        target = current_path if current_path in available_files else available_files[0]
        label = self.file_label_map.get(target, Path(target).name)
        idx = self.cmb_active_file.findText(label)
        if idx >= 0:
            self.cmb_active_file.blockSignals(True)
            self.cmb_active_file.setCurrentIndex(idx)
            self.cmb_active_file.blockSignals(False)
            self._on_active_file_changed()

    # ── graph window ───────────────────────────────────────────────────────

    def open_graph_window(self):
        metric = self.cmb_plot_metric.currentText().strip() or self.last_metric
        file_path = self._selected_active_file_path()
        out = None
        if metric and file_path:
            out = self.metric_results.get(metric, {}).get(file_path)
        if out is None:
            out = self.last_out
            metric = metric or self.last_metric
        if not out or not metric:
            QMessageBox.information(self, "No data", "No analysis results to display.")
            return
        fig = create_metric_plot(metric, out)
        if fig is None:
            QMessageBox.information(self, "No plot", "This metric has no plottable data.")
            return
        label = self.file_label_map.get(file_path, Path(file_path).name) if file_path else "latest"
        dlg = GraphWindow(fig, f"{metric}  ·  {label}", parent=self, style=APP_STYLE)
        dlg.exec()

    # ── metric execution ───────────────────────────────────────────────────

    def _run_metric_worker(self, metric, p, insig, fs, show_plots):
        if metric == "Loudness (ISO 532-1)":
            return Loudness_ISO532_1(insig=insig, fs=fs, field=p["field"], method=p["method"],
                                     time_skip=p["time_skip"], show=show_plots)
        if metric == "Sharpness (DIN 45692)":
            return Sharpness_DIN45692(insig=insig, fs=fs, weight_type=p["weight_type"],
                                      LoudnessField=p["field"], LoudnessMethod=p["method"],
                                      time_skip=p["time_skip"], show_sharpness=show_plots, show_loudness=False)
        if metric == "Roughness (Daniel 1997)":
            return Roughness_Daniel1997(insig=insig, fs=fs, time_skip=p["time_skip"], show=show_plots)
        if metric == "Fluctuation Strength (Osses 2016)":
            return FluctuationStrength_Osses2016(insig=insig, fs=fs, method=p["method_fs"],
                                                  time_skip=p["time_skip"], show=show_plots)
        if metric == "Tonality (Aures 1985)":
            return Tonality_Aures1985(insig=insig, fs=fs, LoudnessField=p["field"],
                                      time_skip=p["time_skip"], show=show_plots)
        if metric == "Annoyance (Di 2016)":
            return PsychoacousticAnnoyance_Di2016(insig=insig, fs=fs, LoudnessField=p["field"],
                                                   time_skip=p["time_skip"], show=show_plots, showPA=show_plots)
        if metric == "Annoyance (Zwicker 1999)":
            return PsychoacousticAnnoyance_Zwicker1999(insig=insig, fs=fs, LoudnessField=p["field"],
                                                        time_skip=p["time_skip"], show=show_plots, showPA=show_plots)
        if metric == "Annoyance (More 2010)":
            return PsychoacousticAnnoyance_More2010(insig=insig, fs=fs, LoudnessField=p["field"],
                                                     time_skip=p["time_skip"], show=show_plots, showPA=show_plots)
        if metric == "EPNL (FAR Part 36)":
            threshold = p.get("threshold_epnl")
            return EPNL_FAR_Part36(insig=insig, fs=fs,
                                   threshold=10.0 if threshold is None else threshold,
                                   show=show_plots)
        if metric == "Loudness (ECMA 418-2)":
            try:
                from metrics_ecma import Loudness_ECMA418_2
            except Exception as e:
                raise RuntimeError(f"ECMA loudness unavailable: {e}")
            return Loudness_ECMA418_2(insig=insig, fs=fs, field=p.get("field",0),
                                      method=p.get("method",1), time_skip=p.get("time_skip",0), show=show_plots)
        if metric == "Roughness (ECMA 418-2)":
            try:
                from metrics_ecma import Roughness_ECMA418_2
            except Exception as e:
                raise RuntimeError(f"ECMA roughness unavailable: {e}")
            return Roughness_ECMA418_2(insig=insig, fs=fs, field=p.get("field",0),
                                       method=p.get("method",1), time_skip=p.get("time_skip",0), show=show_plots)
        if metric == "Tonality (ECMA 418-2)":
            try:
                from metrics_ecma import Tonality_ECMA418_2
            except Exception as e:
                raise RuntimeError(f"ECMA tonality unavailable: {e}")
            out = Tonality_ECMA418_2(insig=insig, fs=fs, field=p.get("field",0),
                                     method=p.get("method",1), time_skip=p.get("time_skip",0), show=show_plots)
            if isinstance(out, dict):
                if "tonalityTDep" in out and "InstantaneousTonality" not in out:
                    out["InstantaneousTonality"] = out["tonalityTDep"]
                if "tonalityAvg" in out and "Kmean" not in out:
                    out["Kmean"] = out["tonalityAvg"]
                if "Kmean" in out and "Tmean" not in out:
                    out["Tmean"] = out["Kmean"]
            return out
        raise ValueError(f"Unknown metric: {metric}")

    def _get_first_scalar(self, val):
        if isinstance(val, (list, tuple)):   return float(val[0])
        if isinstance(val, np.ndarray):      return float(val.ravel()[0])
        try:                                 return float(val)
        except Exception:                    return val

    def _summarize(self, metric: str, out: Dict[str, Any]) -> str:
        lines = []
        units  = {"N":"sone","S":"acum","R":"asper","FS":"vacil","K":"t.u.","T":"t.u.","PA":"annoy"}
        family = {
            "Loudness (ISO 532-1)":"N","Loudness (ECMA 418-2)":"N",
            "Sharpness (DIN 45692)":"S",
            "Roughness (Daniel 1997)":"R","Roughness (ECMA 418-2)":"R",
            "Fluctuation Strength (Osses 2016)":"FS",
            "Tonality (Aures 1985)":"K","Tonality (ECMA 418-2)":"T",
            "Annoyance (Di 2016)":"PA","Annoyance (Zwicker 1999)":"PA","Annoyance (More 2010)":"PA",
        }.get(metric)
        if family:
            for sfx, lbl in [("mean","Mean"),("5","5th%"),("95","95th%"),("max","Max"),("min","Min")]:
                k = f"{family}{sfx}"
                if k in out:
                    lines.append(f"{lbl}: {self._get_first_scalar(out[k]):.3f} {units[family]}")
        if "Loudness" in metric and "Loudness" in out:
            lines.insert(0, f"Loudness: {self._get_first_scalar(out['Loudness']):.3f} sone")
        if "Sharpness" in metric and "Sharpness" in out:
            lines.insert(0, f"Sharpness: {self._get_first_scalar(out['Sharpness']):.3f} acum")
        if "ScalarPA" in out:
            lines.append(f"ScalarPA: {self._get_first_scalar(out['ScalarPA']):.3f} annoy")
        if metric == "EPNL (FAR Part 36)":
            if "EPNL"  in out: lines.append(f"EPNL: {self._get_first_scalar(out['EPNL']):.3f} EPNdB")
            if "PNLM"  in out: lines.append(f"PNLM: {self._get_first_scalar(out['PNLM']):.3f} PNdB")
            if "PNLTM" in out: lines.append(f"PNLTM: {self._get_first_scalar(out['PNLTM']):.3f} PNdB")
        if not lines:
            for tkey, ykey, label in [
                ("time","InstantaneousLoudness","Loudness"),("time","InstantaneousSharpness","Sharpness"),
                ("time","InstantaneousRoughness","Roughness"),("time","InstantaneousFluctuationStrength","Fluctuation"),
                ("time","InstantaneousTonality","Tonality"),("time","InstantaneousPA","Annoyance"),
                ("time","PNL","PNL"),("time","PNLT","PNLT"),
            ]:
                if tkey in out and ykey in out:
                    y = np.asarray(out[ykey]).ravel()
                    lines.append(f"{label} – mean:{np.mean(y):.3f}, max:{np.max(y):.3f}")
                    break
        return "\n".join(lines) if lines else "No summary values."

    # ── figure saving ──────────────────────────────────────────────────────

    def _save_figure(self, metric, out, file_path, split=False):
        fname = Path(file_path).stem
        safe  = metric.replace(" ","_").replace("(","").replace(")","")
        if split:
            figures = create_metric_plot_split(metric, out)
            if isinstance(figures, dict):
                for comp, fig in figures.items():
                    fig.savefig(os.path.join(self.save_dir, f"{safe}_{comp}_{fname}.png"),
                                dpi=100, bbox_inches="tight")
                return
        fig = create_metric_plot(metric, out)
        if fig:
            fig.savefig(os.path.join(self.save_dir, f"{safe}_{fname}.png"),
                        dpi=100, bbox_inches="tight")

    # ── Excel export ───────────────────────────────────────────────────────

    def _export_results_excel(self, results_by_metric):
        wb = Workbook()
        wb.remove(wb.active)
        exported = 0
        total_rows = 0
        total_files: set = set()
        for metric, results in results_by_metric.items():
            if not any(out is not None for _, out, err in results if not err):
                continue
            ws = wb.create_sheet(self._safe_sheet_name(metric, wb.sheetnames))
            row_count, files = self._write_metric_summary_sheet(ws, metric, results)
            if row_count > 0:
                exported += 1
                total_rows += row_count
                total_files.update(files)
            else:
                wb.remove(ws)
        if exported == 0:
            self._append_terminal("\nNo results available for Excel export.")
            QMessageBox.warning(self, "Excel Export", "No data to export.")
            return
        excel_path = os.path.join(self.excel_dir, "metrics_comparison.xlsx")
        wb.save(excel_path)
        QMessageBox.information(self, "Excel Export",
            f"Saved to:\n{excel_path}\n\n{exported} sheet(s), {total_rows} parameters, "
            f"{len(total_files)} file(s).")

    def _write_metric_summary_sheet(self, ws, metric, results):
        ws["A1"] = f"Metric: {metric}"
        ws["A1"].font = XFont(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        collected: Dict[str, Dict] = {}
        all_files: List[str] = []
        for fp, out, err in results:
            if err or not out:
                continue
            fname = Path(fp).stem
            for stat_name, stat_value in self._extract_metric_stats(metric, out).items():
                collected.setdefault(stat_name, {})[fname] = stat_value
            if fname not in all_files:
                all_files.append(fname)
        if not collected:
            return 0, []

        def sort_key(item):
            lbl = item[0]
            if "[PRIMARY]"   in lbl: return (0, lbl)
            if "[SECONDARY]" in lbl: return (1, lbl)
            if "[TERTIARY]"  in lbl: return (2, lbl)
            if "[METADATA]"  in lbl: return (3, lbl)
            return (4, lbl)

        sorted_rows = dict(sorted(collected.items(), key=sort_key))
        ws["A2"] = "Parameter"
        for co, fn in enumerate(all_files, 1):
            ws[f"{get_column_letter(co+1)}2"] = fn
        for ri, (label, vals) in enumerate(sorted_rows.items(), 3):
            display = (label.replace("[PRIMARY] ","").replace("[SECONDARY] ","")
                           .replace("[TERTIARY] ","").replace("[METADATA] ","").replace("[OTHER] ",""))
            ws.cell(row=ri, column=1, value=display)
            for co, fn in enumerate(all_files, 1):
                ws.cell(row=ri, column=co+1, value=vals.get(fn, ""))
        self._format_excel_sheet(ws, len(all_files))
        return len(sorted_rows), all_files

    def _safe_sheet_name(self, metric, existing):
        invalid = '\\/*?:[]'
        name = "".join(c for c in metric if c not in invalid).strip() or "Metric"
        name = name[:31]
        if name not in existing:
            return name
        counter = 2
        while True:
            suffix = f" ({counter})"
            candidate = name[:31-len(suffix)] + suffix
            if candidate not in existing:
                return candidate
            counter += 1

    def _extract_metric_stats(self, metric, out):
        stats = {}
        family_keys = {
            "Loudness":("N","sone"), "Sharpness":("S","acum"),
            "Roughness":("R","asper"), "Fluctuation":("FS","vacil"),
            "Tonality":("K","t.u."), "Annoyance":("PA","annoy"),
        }
        mf = mu = None
        for fam, (prefix, unit) in family_keys.items():
            if fam in metric:
                mf, mu = prefix, unit
                break
        if mf:
            for ks, label in [("5","5th percentile"),("mean","Mean"),("95","95th percentile"),
                               ("max","Max"),("min","Min"),("std","Std Dev")]:
                k = f"{mf}{ks}"
                if k in out:
                    stats[f"[PRIMARY] {label} ({mu})"] = self._get_first_scalar(out[k])
        if "Loudness" in metric:
            if "Loudness"      in out: stats["[PRIMARY] Overall Loudness (sone)"] = self._get_first_scalar(out["Loudness"])
            if "LoudnessLevel" in out: stats["[PRIMARY] Loudness Level (phon)"]   = self._get_first_scalar(out["LoudnessLevel"])
        if "Sharpness" in metric and "Sharpness" in out:
            stats["[PRIMARY] Overall Sharpness (acum)"] = self._get_first_scalar(out["Sharpness"])
        if "Annoyance" in metric:
            if "ScalarPA" in out: stats["[PRIMARY] Scalar PA (annoy)"] = self._get_first_scalar(out["ScalarPA"])
            for wk, wl in [("ws","Sharpness"),("wfr","Roughness/Fluctuation"),("wt","Tonality")]:
                if wk in out:
                    arr = np.asarray(out[wk]).ravel()
                    if arr.size:
                        stats[f"[PRIMARY] Weight {wl} - Mean"] = float(np.mean(arr))
                        stats[f"[PRIMARY] Weight {wl} - Max"]  = float(np.max(arr))
        if "EPNL" in metric:
            if "EPNL"  in out: stats["[PRIMARY] EPNL (EPNdB)"]                 = self._get_first_scalar(out["EPNL"])
            if "PNLM"  in out: stats["[PRIMARY] PNLM - Max PNL (PNdB)"]        = self._get_first_scalar(out["PNLM"])
            if "PNLTM" in out: stats["[PRIMARY] PNLTM - Max Tone-Corr (PNdB)"] = self._get_first_scalar(out["PNLTM"])
        inst_keys = {
            "InstantaneousLoudness":"Loudness","InstantaneousLoudnessLevel":"Loudness Level",
            "InstantaneousSharpness":"Sharpness","InstantaneousRoughness":"Roughness",
            "InstantaneousFluctuationStrength":"Fluctuation Strength",
            "InstantaneousTonality":"Tonality","InstantaneousPA":"Annoyance",
        }
        for key, label in inst_keys.items():
            if key in out:
                arr = np.asarray(out[key]).ravel()
                if arr.size > 1:
                    stats[f"[SECONDARY] {label} - Mean"] = float(np.mean(arr))
                    stats[f"[SECONDARY] {label} - Max"]  = float(np.max(arr))
                    stats[f"[SECONDARY] {label} - Min"]  = float(np.min(arr))
                    stats[f"[SECONDARY] {label} - Std"]  = float(np.std(arr))
        if "SpecificLoudness" in out:
            arr = np.asarray(out["SpecificLoudness"]).ravel()
            if arr.size:
                stats["[TERTIARY] Specific Loudness - Mean"] = float(np.mean(arr))
                stats["[TERTIARY] Specific Loudness - Max"]  = float(np.max(arr))
        for key in out:
            if key not in stats:
                val = out[key]
                if isinstance(val, (int, float)):
                    stats[f"[OTHER] {key}"] = val
                elif isinstance(val, (list, np.ndarray)):
                    arr = np.asarray(val).ravel()
                    if 1 <= arr.size <= 5:
                        stats[f"[OTHER] {key}"] = float(arr[0]) if arr.size == 1 else f"array({arr.size})"
        return stats

    def _format_excel_sheet(self, ws, num_files=1):
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = XFont(bold=True, color="FFFFFF", size=11)
        pri_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        pri_font = XFont(bold=True, size=10)
        sec_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        sec_font = XFont(size=9)
        ter_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ter_font = XFont(size=9, italic=True)
        meta_font = XFont(size=8, color="666666")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        for cell in ws[2]:
            if cell.value:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = XAlignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
        for ri, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row,
                                               min_col=1, max_col=ws.max_column), 3):
            rl = ws.cell(row=ri, column=1).value or ""
            if   "[PRIMARY]"   in rl: rf, lf = pri_fill, pri_font
            elif "[SECONDARY]" in rl: rf, lf = sec_fill, sec_font
            elif "[TERTIARY]"  in rl: rf, lf = ter_fill, ter_font
            else:                     rf, lf = None, meta_font
            for ci, cell in enumerate(row, 1):
                cell.border = border
                if ci == 1:
                    cell.font = lf
                    if rf: cell.fill = rf
                    cell.alignment = XAlignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"
                    cell.alignment = XAlignment(horizontal="center", vertical="center")
                    if rf: cell.fill = rf
        ws.column_dimensions["A"].width = 50
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18


# ── entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    palette = QPalette()
    cr = QPalette.ColorRole
    palette.setColor(cr.Window,          QColor(C["bg"]))
    palette.setColor(cr.WindowText,      QColor(C["text"]))
    palette.setColor(cr.Base,            QColor(C["surface"]))
    palette.setColor(cr.AlternateBase,   QColor(C["card"]))
    palette.setColor(cr.ToolTipBase,     QColor(C["card"]))
    palette.setColor(cr.ToolTipText,     QColor(C["text"]))
    palette.setColor(cr.Text,            QColor(C["text"]))
    palette.setColor(cr.Button,          QColor(C["card"]))
    palette.setColor(cr.ButtonText,      QColor(C["text"]))
    palette.setColor(cr.BrightText,      QColor("#ffffff"))
    palette.setColor(cr.Highlight,       QColor(C["accent"]))
    palette.setColor(cr.HighlightedText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorGroup.Disabled, cr.Text,       QColor(C["text3"]))
    palette.setColor(QPalette.ColorGroup.Disabled, cr.ButtonText, QColor(C["text3"]))
    app.setPalette(palette)

    window = SQATModernApp()
    window.show()
    sys.exit(app.exec())
